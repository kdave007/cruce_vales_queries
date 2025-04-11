from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

class ExcelService:
    def __init__(self, file_name):
        """
        init
        """
        self.file_name = file_name
        self.workbook = Workbook()
        # Remove default sheet that comes with new workbook
        self.workbook.remove(self.workbook.active)

    def create_sheet(self, sheet_name):
        """
        Create a new sheet with the given name
        Returns the created sheet
        """  
        return self.workbook.create_sheet(sheet_name)

    def write_headers(self, sheet, headers):
        """
        Write and format headers in the given sheet
        Args:
            sheet: Worksheet to write to
            headers: List of header names
        """ 
        # Style for headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(fill_type="solid", fgColor="366092")
        center_align = Alignment(horizontal="center")

        for col, header in enumerate (headers,1):
            cell = sheet.cell(row=1,column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align


    def write_data(self, sheet, data):
        """
        Write data rows to the given sheet
        Args:
            sheet: Worksheet to write to
            data: List of dictionaries containing the data
        """
        # Start from row 2 (row 1 has headers)
        for row_index, row_data in enumerate(data,2):
            for col_index, value in enumerate(row_data.values(),1):
                #insert 
                cell = sheet.cell(row=row_index, column=col_index)
                cell.value = value
                cell.alignment = Alignment(horizontal="center")



    def save(self):
        """
        Save workbook to file
        """
        self.workbook.save(self.file_name)            